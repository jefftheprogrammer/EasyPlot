# EasyPlot
# By Mihai Stefan Merlas
# Allows user to easily make plots
#############################

import matplotlib.pyplot as plt
import openpyxl as excel
import numpy as np


class Variable:
    def __init__(self):
        self.var_marker = None
        self.type = None
        self.name = None
        self.unit = None
        self.data = None

    def set_variable(self, var_marker, type, name, unit, data):
        self.var_marker = var_marker
        self.type = type
        self.name = name
        self.unit = unit
        self.data = data

    @staticmethod
    def get_data(self):
        return self.data


class DataFromFile:

    @staticmethod
    def find_file():  # loads file
        filename = (input("enter filename") + ".xlsx")
        wb = excel.load_workbook(filename=filename)
        sheet = wb.active
        return sheet

    @staticmethod
    def extract_columns(sheet):
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        obj_list = []
        data_from_sheet = []

        for column in sheet.iter_cols(min_row=1,
                                      max_row=num_rows,
                                      min_col=1,
                                      max_col=num_cols):
            data_from_sheet.append(column)

        for col in range(num_cols):
            var_marker = data_from_sheet[col][0].value
            type = data_from_sheet[col][1].value
            name = data_from_sheet[col][2].value
            unit = data_from_sheet[col][3].value
            data = []
            for row in range(4, num_rows):
                data.append(data_from_sheet[col][row].value)

            obj_list.append(Variable())
            obj_list[len(obj_list) - 1].set_variable(var_marker, type, name, unit, data)

        return obj_list, num_cols

    @staticmethod
    def no_of_elements(dictionaries):
        s = 0
        for key in dictionaries.keys():
            s += len(dictionaries[key])

        return s

    def get_data(self):
        sheet = DataFromFile.find_file()
        data, num_cols = DataFromFile.extract_columns(sheet)
        print(data[0].name, data[0].type, data[0].data)
        new_group = {}
        new_data = {}

        #  adding y variables to new_data dictionary
        for variable in data:
            if variable.var_marker == "y":
                if variable.type == "DATA":
                    new_group["DATA"] = variable
                elif variable.type == "ERROR":
                    new_group["ERROR"] = variable

        new_data["y"] = new_group
        new_group = {}
        a = 1
        done = False

        # adding x variables to new_data dictionary
        while done == False:
            for b in range(len(data)):
                total_var = self.no_of_elements(new_data)  # nos of elements in dict
                if data[b].var_marker == "x{}".format(a) and num_cols > total_var:
                    if data[b].type == "DATA":
                        new_group["DATA"] = data[b]
                    elif data[b].type == "ERROR":
                        new_group["ERROR"] = data[b]
                    print(new_group)
                elif num_cols <= total_var + len(new_group):
                    done = True
            if len(new_group) > 0:
                new_data["x{}".format(a)] = new_group
            new_group = {}
            a += 1

        return new_data


class Graph:
    def __init__(self):
        self.data = None

    @staticmethod
    def choose_marker():
        print(""" .:point
,:pixel
o:circle
v:triangle down
^:triangle up
<:triangle left
>:triangle right
1:tri down
2:tri up
3:tri left
4:tri right
p:plus (thick)
+:plus
X:X (thick)
x:X
D:diamond (thick)
d:diamond
|:vertical line
_:horizontal line
""")

    @staticmethod
    def check(data):
        sameness_unit = True
        sameness_variable_name = True
        sep_subplots_req = False
        if len(data) > 2:
            for a in range(1, len(data) - 1):
                if data["x1"]["DATA"].unit != data["x{}".format(a)]["DATA"].unit:
                    sameness_unit = False
                if data["x1"]["DATA"].name != data["x{}".format(a)]["DATA"].name:
                    sameness_name = False

        if sameness_unit == False or sameness_variable_name == False:
            sep_subplots_req = True

        return sep_subplots_req

    @staticmethod
    def choose_setting(marker_name):
        print("please choose setting for {}".format(marker_name))
        setting = {}
        setting["marker"] = input("please choose marker")
        setting["color"] = input("please choose color")
        setting["capsize"] = 1.0
        return setting

    def plot_same_sbplt(self):
        plt.xlabel("{} ({})".format(self.data["x1"]["DATA"].name, self.data["x1"]["DATA"].unit))
        plt.ylabel("{} ({})".format(self.data["y"]["DATA"].name, self.data["y"]["DATA"].unit))

        for a in range(1, len(self.data)):
            marker = "x{}".format(a)
            setting = Graph.choose_setting(marker)
            plt.errorbar(self.data[marker]["DATA"].data, self.data["y"]["DATA"].data,
                         xerr=self.data[marker]["ERROR"].data, yerr=self.data["y"]["ERROR"].data,
                         marker=setting["marker"], color=setting["color"], capsize=setting["capsize"])

        plt.show()

    def plot_diff_sbplt(self):
        objs = []
        print("You have {} x variables, please type in the configuration".format(len(self.data) - 1))

        for a in range(1, len(self.data) - 1):
            nrows = int(input("nrows:"))
            ncols = int(input("ncols:"))
            index = int(input("index"))
            marker = "x{}".format(a)
            objs.append(plt.subplot(nrows, ncols, index))
            setting = Graph.choose_setting(marker)
            objs[a - 1].errorbar(self.data[marker]["DATA"].data, self.data["y"]["DATA"].data,
                                 xerr=data[marker]["ERROR"].data, yerr=data["y"]["ERROR"].data,
                                 marker=setting["marker"], color=setting["color"], capsize=setting["capsize"])

        plt.show()

    def scatterPlot(self):
        D = DataFromFile()
        self.data = D.get_data()
        sep_subplots = Graph.check(self.data)
        if sep_subplots == True:
            self.plot_diff_sbplt()
        elif sep_subplots == False:
            self.plot_same_sbplt()


G = Graph()
G.scatterPlot()
